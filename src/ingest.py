"""
Document ingestion module for Open Brain.
Supports: PDF, images (JPG/PNG/WEBP/GIF), Word (.docx), Excel (.xlsx/.xls), plain text.
Uses vision model for images and scanned PDFs, text extraction for everything else.
"""
import os
import io
import json
from typing import Optional

# PDF
try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

# Word
try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

# Excel
try:
    import openpyxl
except ImportError:
    openpyxl = None

# Images
try:
    from PIL import Image
except ImportError:
    Image = None


MIME_MAP = {
    ".pdf": "application/pdf",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
    ".tiff": "image/tiff",
    ".tif": "image/tiff",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".doc": "application/msword",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".txt": "text/plain",
    ".md": "text/markdown",
    ".csv": "text/csv",
}

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff", ".tif"}


def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract text from a PDF. Returns empty string if the PDF is image-only (scanned)."""
    if PdfReader is None:
        return "[PyPDF2 not installed]"
    reader = PdfReader(io.BytesIO(file_bytes))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"--- Page {i+1} ---\n{text.strip()}")
    return "\n\n".join(pages)


def extract_text_from_docx(file_bytes: bytes) -> str:
    """Extract text from a Word .docx file."""
    if DocxDocument is None:
        return "[python-docx not installed]"
    doc = DocxDocument(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def extract_text_from_excel(file_bytes: bytes) -> str:
    """Extract text from an Excel file, sheet by sheet."""
    if openpyxl is None:
        return "[openpyxl not installed]"
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            sheets.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheets)


def pdf_to_images(file_bytes: bytes) -> list[bytes]:
    """Convert PDF pages to images for vision OCR using PyMuPDF (no system deps)."""
    try:
        import fitz  # pymupdf
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        result = []
        for page in doc:
            pix = page.get_pixmap(dpi=200)
            result.append(pix.tobytes("png"))
        doc.close()
        return result
    except ImportError:
        return []
    except Exception as e:
        print(f"pdf_to_images error: {e}", flush=True)
        return []


def ingest_document(filename: str, file_bytes: bytes) -> dict:
    """
    Main ingestion entry point.
    Returns: {"text": str, "source_type": str, "method": str, "filename": str}
    """
    ext = os.path.splitext(filename)[1].lower()

    # Images → vision model directly
    if ext in IMAGE_EXTENSIONS:
        from llm import describe_image
        mime = MIME_MAP.get(ext, "image/png")
        text = describe_image(file_bytes, mime_type=mime)
        return {
            "text": text,
            "source_type": "document_image",
            "method": "vision_ocr",
            "filename": filename,
        }

    # PDF → try text extraction first, fall back to vision OCR
    if ext == ".pdf":
        text = extract_text_from_pdf(file_bytes)
        if len(text.strip()) > 50:
            return {
                "text": text,
                "source_type": "document_pdf",
                "method": "text_extraction",
                "filename": filename,
            }
        # Scanned PDF — try vision OCR on pages
        from llm import describe_image
        page_images = pdf_to_images(file_bytes)
        if page_images:
            pages = []
            for i, img_bytes in enumerate(page_images):
                page_text = describe_image(img_bytes, mime_type="image/png")
                pages.append(f"--- Page {i+1} ---\n{page_text}")
            return {
                "text": "\n\n".join(pages),
                "source_type": "document_pdf",
                "method": "vision_ocr",
                "filename": filename,
            }
        return {
            "text": text or "[Could not extract text from this PDF — it may be scanned. Install pdf2image + poppler for OCR support.]",
            "source_type": "document_pdf",
            "method": "text_extraction",
            "filename": filename,
        }

    # Word
    if ext in (".docx", ".doc"):
        text = extract_text_from_docx(file_bytes)
        return {
            "text": text,
            "source_type": "document_word",
            "method": "text_extraction",
            "filename": filename,
        }

    # Excel
    if ext in (".xlsx", ".xls"):
        text = extract_text_from_excel(file_bytes)
        return {
            "text": text,
            "source_type": "document_excel",
            "method": "text_extraction",
            "filename": filename,
        }

    # Plain text / CSV / Markdown
    if ext in (".txt", ".md", ".csv"):
        text = file_bytes.decode("utf-8", errors="replace")
        return {
            "text": text,
            "source_type": f"document_{ext.lstrip('.')}",
            "method": "text_extraction",
            "filename": filename,
        }

    # Unknown — try as text
    try:
        text = file_bytes.decode("utf-8", errors="replace")
        return {
            "text": text,
            "source_type": "document_unknown",
            "method": "text_extraction",
            "filename": filename,
        }
    except Exception:
        return {
            "text": f"[Unsupported file type: {ext}]",
            "source_type": "document_unknown",
            "method": "none",
            "filename": filename,
        }
